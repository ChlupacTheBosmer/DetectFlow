from typing import Dict, Callable, Optional, List, Dict, Any, Union, Type
import os
import logging
import pandas as pd
import re
try:
    import openpyxl
    import xlwings as xw
    excel_available = True
except ImportError:
    excel_available = False

class BaseAnnotationFile:
    def __init__(self, filepath, output_path: str = None):

        if not excel_available:
            raise ImportError("Please run 'pip install detectflow[excel]' to install the DetectFlow package with Excel handling libraries.")

        # Define variables
        self.filepath = filepath
        self.output_path = output_path if output_path and os.path.isfile(output_path) else os.path.join(output_path, os.path.basename(filepath)) if output_path else None

    @staticmethod
    def _evaluate_string_formula(cell):
        if isinstance(cell, (int, float)):
            return cell
        elif str(cell).startswith('='):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'].value = cell
            value = ws['A1'].value
            wb.close()
            return value
        else:
            return cell

    def _convert_months(self, cell):
        cell = self._evaluate_string_formula(cell)
        months_short = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        months_long = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06',
                       'July': '07', 'August': '08', 'September': '09', 'October': '10', 'November': '11',
                       'December': '12'}
        months_number = {str(i): f'{i:02d}' for i in range(1, 13)}

        if cell in months_short:
            return months_short[cell]
        elif cell in months_long:
            return months_long[cell]
        elif str(cell) in months_number:
            return months_number[str(cell)]
        else:
            if not cell == "":
                logging.warning(f"Invalid month abbreviation in Excel file: {cell}")
            return '01'

    def _convert_year(self, cell):
        try:
            return int(self._evaluate_string_formula(cell))
        except ValueError as e:
            if not cell == "":
                logging.warning(f"Unexpected value in Excel file. Error: {e}, value: '{cell}'")
            return 2000 if cell != "" else ""

    def _convert_time_data(self, cell):
        try:
            return "{:02d}".format(int(self._evaluate_string_formula(cell)))
        except ValueError as e:
            if not cell == "":
                logging.warning(f"Unexpected value in Excel file. Error: {e}, value: '{cell}'")
            return "00" if cell != "" else ""

    def _convert_bool(self, cell):
        cell = self._evaluate_string_formula(cell)
        return cell == 1

    def _construct_timestamp(self, dataframe):
        dataframe = dataframe.copy()
        dataframe['ts'] = dataframe.iloc[:, 0:6].apply(lambda x: f"{x.iloc[0]}{x.iloc[1]}{x.iloc[2]}_{x.iloc[3]}_{x.iloc[4]}_{x.iloc[5]}", axis=1)
        return dataframe

    def _save_temp_file(self, dataframe, output_path: str = None):

        # If the output path is not specified, use the default one
        output_path = output_path if output_path and os.path.isfile(output_path) else self.output_path

        if not output_path:
            logging.error("Output path is not specified. Cannot save the temporary file.")
            return

        # Save temp file
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        dataframe.to_excel(output_path, index=False)

    def _load_excel_file(self, cols, converters):
        try:
            # Read the Excel file, skipping the first two rows
            dataframe = pd.read_excel(self.filepath, usecols=cols, skiprows=2, header=None, converters=converters)
        except ValueError as e:
            import tempfile

            logging.error(f"Error reading Excel file {self.filepath}. Error message: {e}")

            # Open the Excel workbook using xlwings
            workbook = xw.Book(self.filepath)
            sheet = workbook.sheets[0]

            # Remove any filters that might be present
            if sheet.api.AutoFilterMode:
                sheet.api.AutoFilterMode = False

            # Save to temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_filepath = temp_file.name

            workbook.save(temp_filepath)
            workbook.close()

            # Read with pandas again
            try:
                dataframe = pd.read_excel(temp_filepath, usecols=cols, skiprows=2, header=None, converters=converters)
            except ValueError as e:
                logging.error(f"Attempt to fix errors in Excel file {self.filepath} failed. Error message: {e}")
                return None
            logging.info(f"Removed filters from Excel file {self.filepath}. Saved a copy to {temp_filepath}")

        return dataframe


class AnnotationFile(BaseAnnotationFile):
    def __init__(self, filepath, **kwargs):

        #super Init superclass
        super().__init__(filepath)

        # Define variables
        self._read_file_method = self.read_from_sqlite if filepath.endswith('.db') else self.read_from_excel
        self.filepath = filepath
        self._load_time_data = kwargs.get('load_time_data', True)
        self._load_visit_data = kwargs.get('load_visit_data', True)
        self._load_visitor_data = kwargs.get('load_visitor_data', True)
        self._load_behavior_data = kwargs.get('load_behavior_data', True)
        self._table_name = kwargs.get('table_name', None)
        self._dataframe = None

    @property
    def dataframe(self):
        if self._dataframe is None:
            try:
                self._dataframe = self._read_file_method()
            except Exception as e:
                logging.error(f"Error reading DataFrame from '{self.filepath}'. Error: {e}")
                self._dataframe = None
        return self._dataframe

    @dataframe.setter
    def dataframe(self, value: pd.DataFrame):
        self._dataframe = value

    def read_from_excel(self) -> Optional[pd.DataFrame]:

        # The default values of cols to be extracted are:
        # 0 - A - Year
        # 1 - B - Month - !
        # ...
        # 5 - F - Seconds
        # 15 - P - Visitor arrival - filter column
        # 18 - S - Visit duration in seconds
        # 19 - T - Time of departure - Hours
        # 20 - U - Time of departure - Minutes
        # 21 - V - Time of departure - Seconds
        # 23 - X - Insect species
        # 24 - Y - Insect Order

        # For more info see the example excel table in resources/exc there I wrote down the numbers of columns, datatypes etc.

        # Define the number of cols for each data extraction module
        time_cols: list[int] = [0, 1, 2, 3, 4, 5]
        visit_cols: list[int] = [15, 18, 19, 20, 21]
        visitor_cols: list[int] = [23, 24]
        behavior_cols: list[int] = [27, 34, 36, 40, 46, 47, 48, 49]

        # Define convertors for each data extraction module
        time_converters: Dict[int, Callable] = {0: self._convert_year,
                                                1: self._convert_months,
                                                2: self._convert_time_data,
                                                3: self._convert_time_data,
                                                4: self._convert_time_data,
                                                5: self._convert_time_data}

        visit_converters: Dict[int, Callable] = {15: self._convert_bool,
                                                 18: self._evaluate_string_formula,
                                                 19: self._convert_time_data,
                                                 20: self._convert_time_data,
                                                 21: self._convert_time_data}

        visitor_converters: Dict[int, Callable] = {23: self._evaluate_string_formula,
                                                   24: self._evaluate_string_formula}

        behavior_converters: Dict[int, Callable] = {27: self._evaluate_string_formula,
                                                    34: self._convert_bool,
                                                    36: self._convert_bool,
                                                    40: self._convert_bool,
                                                    46: self._convert_bool,
                                                    47: self._evaluate_string_formula,
                                                    48: self._convert_bool,
                                                    49: self._evaluate_string_formula}

        modules = [(self._load_time_data, time_cols, time_converters),
                   (self._load_visit_data, visit_cols, visit_converters),
                   (self._load_visitor_data, visitor_cols, visitor_converters),
                   (self._load_behavior_data, behavior_cols, behavior_converters)]

        # Assemble the cols and converters to be used in reading the Excel file
        cols = []
        converters = {}
        for module_bool, module_cols, module_converters in modules:
            if module_bool:
                cols += module_cols
                converters.update(module_converters)

        # Open the Excel, resolve any issues and load dataframe
        dataframe = self._load_excel_file(cols, converters)

        # Mapping dictionary for column renaming
        column_mapping = {
            0: 'year',
            1: 'month',
            2: 'day',
            3: 'hour_a',
            4: 'min_a',
            5: 'sec_a',
            15: 'visit',
            18: 'duration',
            19: 'hour_d',
            20: 'min_d',
            21: 'sec_d',
            23: 'vis_id',
            24: 'vis_ord',
            27: 'no_flo',
            34: 'f_pol',
            36: 'f_nec',
            40: 'f_fp',
            46: 'con_m',
            47: 'no_con_m',
            48: 'con_f',
            49: 'no_con_f'
        }

        # Rename columns using the mapping dictionary
        dataframe.rename(columns=column_mapping, inplace=True)

        # Filter data frame based on whether the value in the column of index 6 (P - visitor arrival) is 1.
        filtered_dataframe = dataframe[dataframe['visit'] == True]

        # Add another column called "ts" for timestamp
        filtered_dataframe = self._construct_timestamp(filtered_dataframe)

        # Preprocess the visitor descriptions if applicable
        if self._load_visitor_data:
            # Check which column has fewer NAs
            chosen_column, other_column = ('vis_id', 'vis_ord') if filtered_dataframe['vis_id'].isna().sum() <= filtered_dataframe['vis_ord'].isna().sum() else ('vis_ord', 'vis_id')

            # Replace NAs in chosen_column with values from other_column if they are not NAs
            filtered_dataframe = filtered_dataframe.copy()
            filtered_dataframe.fillna({chosen_column: filtered_dataframe[other_column]}, inplace=True)
            filtered_dataframe[other_column] = filtered_dataframe[chosen_column]

        # Save temporary file from dataframe to excel
        if self.output_path:
            self._save_temp_file(filtered_dataframe)

        # Return the dataframe
        return filtered_dataframe

    def read_from_sqlite(self, table_name: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        Read the dataframe from a SQLite database. Does not perform any post-processing correct format is assumed.

        :param table_name: Optional[str]: name of the table to read from
        :return: dataframe: pd.DataFrame: dataframe read from the SQLite database
        """
        from detectflow.manipulators.database_manipulator import DatabaseManipulator

        # Define the database manipulator
        db = DatabaseManipulator(self.filepath)

        # Get tables from the database
        tables = db.get_table_names()

        # Check if the table exists
        table_name = table_name if table_name else self._table_name
        if table_name:
            if table_name not in tables:
                raise ValueError(f"Table '{table_name}' not found in the database.")
        else:
            try:
                table_name = db.get_table_names()[0]
            except IndexError:
                logging.error("No tables found in the database.")
                return None

        # Read the dataframe from the database using pandas custom method
        try:
            conn = db.create_connection()
            dataframe = pd.read_sql(f"SELECT * FROM {table_name}", conn)
        except Exception as e:
            logging.error(f"Error reading DataFrame from '{table_name}' table in '{self.filepath}' database. Error: {e}")
            dataframe = None
        finally:
            db.close_connection()

        return dataframe

    def save_to_sqlite(self, db_path: Optional[str] = None, table_name: Optional[str] = None) -> Optional[str]:
        from detectflow.manipulators.database_manipulator import DatabaseManipulator

        # Get variables
        db_path = db_path if db_path else os.path.join(self.output_path if self.output_path else os.path.dirname(self.filepath), f"{os.path.splitext(os.path.basename(self.filepath))[0]}.db")
        table_name = table_name if table_name else self._table_name if self._table_name else "data"

        # Define the database manipulator
        db = DatabaseManipulator(db_path)

        try:
            # Create the table
            conn = db.create_connection()
            self.dataframe.to_sql(table_name, conn, if_exists='replace', index=False)
        except Exception as e:
            logging.error(f"Error saving DataFrame to '{table_name}' table in '{db_path}' database. Error: {e}")
            db_path = None
        finally:
            db.close_connection()

        return db_path

class CustomAnnotationFile(BaseAnnotationFile):
        def __init__(self, filepath):
            super().__init__(filepath)
            self._dataframe = None

        @property
        def dataframe(self):
            if self._dataframe is None:
                self._dataframe = self.read_from_excel()
            return self._dataframe

        def read_from_excel(self):
            # Define the columns to extract and the corresponding converters
            cols = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            converters = {0: self._convert_year, 1: self._convert_months, 2: self._convert_time_data,
                          3: self._convert_time_data, 4: self._convert_time_data, 5: self._convert_time_data,
                          6: self._evaluate_string_formula, 7: self._convert_time_data, 8: self._convert_time_data,
                          9: self._convert_time_data}

            # Read the Excel file, skipping the first two rows - follow the custom format
            dataframe = self._load_excel_file(cols, converters)

            # Make a copy to solve the error when trying to access only a slice
            dataframe = dataframe.copy()

            # Mapping dictionary for column renaming
            column_mapping = {
                0: 'year',
                1: 'month',
                2: 'day',
                3: 'hour_a',
                4: 'min_a',
                5: 'sec_a',
                6: 'duration',
                7: 'hour_d',
                8: 'min_d',
                9: 'sec_d'
            }

            # Rename columns using the mapping dictionary
            dataframe.rename(columns=column_mapping, inplace=True)

            # Add another column called "ts" for timestamp
            dataframe = self._construct_timestamp(dataframe)

            # Save temporary file from dataframe to excel
            if self.output_path:
                self._save_temp_file(dataframe)

            # Return the dataframe
            return dataframe


class AnnotationFileDiagnoser:
    def __init__(self, filepath: str, video_paths: Optional[List[str]] = None):
        self.filepath = filepath
        self.parent_dir = os.path.dirname(filepath)
        self.annotation_file = None
        self.video_paths = video_paths if video_paths else []
        self._report = None

    @property
    def report(self):
        if self._report is None:
            self._report = {**self.check_excel_file(), **self.check_video_alignment(self.video_paths)}
        return self._report

    def check_excel_file(self) -> Dict[str, Any]:
        report = {}
        try:
            self.annotation_file = AnnotationFile(self.filepath)
            df = self.annotation_file.dataframe
            report['excel_status'] = 'success'
            report['dataframe_diagnosis'] = self.diagnose_dataframe(df)
        except Exception as e:
            error_msg = str(e)
            print(error_msg)
            report['excel_status'] = 'error'
            report['error_message'] = self.diagnose_error(error_msg)
        return report

    def check_video_alignment(self, video_paths: List[str]) -> Dict[str, Any]:
        report = {}
        try:
            annotation_list = get_annotations_from_dataframe_with_videos(self.annotation_file.dataframe, video_paths)
            if not isinstance(annotation_list, AnnotationList) or len(annotation_list) == 0:
                raise ValueError("No relevant video paths left.")
            report['video_alignment_status'] = 'success'
            report['relevant_videos'] = annotation_list.video_paths
        except Exception as e:
            error_msg = str(e)
            report['video_alignment_status'] = 'error'
            report['error_message'] = self.diagnose_error(error_msg)
        return report

    @staticmethod
    def diagnose_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
        diagnosis = {
            "shape": df.shape,
            "head": df.head().to_dict(),
            "describe": df.describe().to_dict(),
            "non_null_counts": df.count().to_dict(),
            "unique_counts": {col: df[col].nunique() for col in df.columns},
            "nan_counts": {col: df[col].isna().sum() for col in df.columns}
        }
        return diagnosis

    @staticmethod
    def diagnose_error(error_msg: str) -> str:
        error_patterns = {
            "interactive mode of xlwings": (
                "There is a filter applied to the data in the workbook. This can be fixed automatically on "
                "Windows or macOS but not here. Please remove the filter manually."
            ),
            "list index out of range": (
                "The workbook layout is incorrect. Copy cells from AB:1 to AX:2 from a healthy workbook "
                "to the same range in the problematic workbook."
            ),
            "usecols with out-of-bounds indices": (
                "The workbook layout is incorrect. Copy cells from AB:1 to AX:2 from a healthy workbook "
                "to the same range in the problematic workbook."
            ),
            "has no attribute": (
                "The workbook layout is incorrect. Copy cells from AB:1 to AX:2 from a healthy workbook "
                "to the same range in the problematic workbook."
            ),
            "No relevant video paths left": (
                "Check if the Excel name matches the video names. Ensure that the recording and Excel workbook "
                "are not mixed up and that sorting is case-specific."
            ),
        }

        for pattern, message in error_patterns.items():
            if re.search(pattern, error_msg, re.IGNORECASE):
                return message
        return (
            "An unspecified error occurred. Check the diagnostics report and ensure that:\n"
            "- Videos have the correct naming convention\n"
            "- Dates correspond between Excel and videos\n"
            "- Formulas are not broken\n"
            "- The workbook has no filter applied\n"
            "- The video folder and Excel file are named identically\n"
            "- You have a matching Excel and recording folder."
        )


from datetime import datetime, timedelta
from typing import Type, Tuple
from detectflow.utils.name import parse_recording_name
from detectflow.video.video_data import get_video_file_instance


class Annotation:
    def __init__(self,
                 start_frame: Optional[int] = None,
                 end_frame: Optional[int] = None,
                 video_path: Optional[str] = None,
                 start_time: Optional[datetime] = None,
                 end_time: Optional[datetime] = None,
                 start_video_time: Optional[timedelta] = None,
                 end_video_time: Optional[timedelta] = None,
                 duration: Optional[timedelta] = None,
                 video_id: Optional[str] = None,
                 recording_id: Optional[str] = None,
                 video_fps: Optional[int] = None,
                 detection_results: Optional[List[Type["DetectionResults"]]] = None,
                 flowers: Optional[int] = None,
                 flowers_boxes: Optional[Type["DetectionBoxes"]] = None,
                 flowers_visited: Optional[int] = None):

        # Set public attrs
        self.video_path = video_path
        self.detection_results = detection_results

        # Set private attrs
        self._start_frame = start_frame
        self._end_frame = end_frame
        self._start_time = start_time
        self._end_time = end_time
        self._start_video_time = start_video_time
        self._end_video_time = end_video_time
        self._duration = duration
        self._video_id = video_id
        self._recording_id = recording_id
        self._video_fps = video_fps
        self._flowers = flowers
        self._flowers_boxes = flowers_boxes
        self._flowers_visited = flowers_visited

    @classmethod
    def from_timestamp(cls, timestamp: str, duration: Union[int, timedelta], **kwargs):
        duration = duration if isinstance(duration, timedelta) else timedelta(seconds=duration)
        start_time = datetime.strptime(timestamp, "%Y%m%d_%H_%M_%S")
        end_time = start_time + duration
        return cls(start_time=start_time, end_time=end_time, duration=duration, **kwargs)

    @classmethod
    def from_detection_results(cls, detection_results: Union[List[Type["DetectionResults"]], Tuple[Type["DetectionResults"], ...]], **kwargs):

        # Get the start and end frames from the detection results
        try:
            start_frame = min([result.frame_number for result in detection_results])
            end_frame = max([result.frame_number for result in detection_results])
        except AttributeError:
            logging.error("Detection results do not have frame numbers. Cannot calculate start and end frames.")
            start_frame = None
            end_frame = None

        video_path = None
        if detection_results and len(detection_results) > 0:
            if detection_results[0].source_type == 'video':
                video_path = detection_results[0].source_path

        return cls(start_frame=start_frame, end_frame=end_frame, detection_results=detection_results, video_path=video_path, **kwargs)

    @property
    def start_frame(self):
        if self._start_frame is None:
            self._start_frame = self._get_start_frame()
        return self._start_frame

    @start_frame.setter
    def start_frame(self, value: int):
        self._start_frame = value

    def _get_start_frame(self):
        # Get the start frame from the video file
        start_frame = None
        if self.start_video_time:
            try:
                start_frame = int(self.start_video_time.total_seconds() * self.video_fps)
            except TypeError:
                logging.error("Video FPS is not set. Cannot calculate start frame.")
        elif self.start_time and self.video_path and os.path.isfile(self.video_path):
            try:
                video = get_video_file_instance(self.video_path)
                start_frame = int((self.start_time - video.start_time).total_seconds() * self.video_fps)
            except Exception as e:
                logging.error(f"Failed to get start frame: {e}")
        else:
            logging.error("Start video time and video path are not set. Cannot calculate start frame.")

        return start_frame

    @property
    def end_frame(self):
        if self._end_frame is None:
            self._end_frame = self._get_end_frame()
        return self._end_frame

    @end_frame.setter
    def end_frame(self, value: int):
        self._end_frame = value

    def _get_end_frame(self):
        # Get the end frame from the video file
        end_frame = None
        if self.end_video_time:
            try:
                end_frame = int(self.end_video_time.total_seconds() * self.video_fps)
            except TypeError:
                logging.error("Video FPS is not set. Cannot calculate end frame.")
        elif self.duration and self.start_frame is not None:
            try:
                end_frame = int(self.start_frame + self.duration.total_seconds() * self.video_fps)
            except TypeError:
                logging.error("Video FPS is not set. Cannot calculate end frame.")
        elif self.start_time and self.duration and self.video_path and os.path.isfile(self.video_path):
            try:
                video = get_video_file_instance(self.video_path)
                end_frame = int((self.start_time + self.duration - video.start_time).total_seconds() * self.video_fps)
            except Exception as e:
                logging.error(f"Failed to get end frame: {e}")
        else:
            logging.error("End video time or duration are not set. Cannot calculate end frame.")

        return end_frame

    @property
    def start_video_time(self):
        if self._start_video_time is None:
            self._start_video_time = self._get_start_video_time()
        return self._start_video_time

    @start_video_time.setter
    def start_video_time(self, value: timedelta):
        self._start_video_time = value

    def _get_start_video_time(self):
        # Get the start video time from the video file
        if self._start_frame is not None:
            try:
                start_video_time = timedelta(seconds=self._start_frame / self.video_fps)
            except TypeError:
                logging.error("FPS is not set. Cannot calculate start video time.")
                start_video_time = None
        else:
            logging.error("Start frame is not set. Cannot calculate start video time.")
            start_video_time = None

        return start_video_time

    @property
    def end_video_time(self):
        if not self._end_video_time:
            self._end_video_time = self._get_end_video_time()
        return self._end_video_time

    @end_video_time.setter
    def end_video_time(self, value: timedelta):
        self._end_video_time = value

    def _get_end_video_time(self):
        # Get the end video time from the video file
        if self._end_frame is not None:
            try:
                end_video_time = timedelta(seconds=self._end_frame / self.video_fps)
            except TypeError:
                logging.error("FPS is not set. Cannot calculate end video time.")
                end_video_time = None
        elif self.duration and self.start_video_time is not None:
            try:
                end_video_time = self.start_video_time + self.duration
            except TypeError:
                logging.error("Start video time or duration have incorrect format. Cannot calculate end video time.")
                end_video_time = None
        else:
            logging.error("End frame, start video time, and duration are not set. Cannot calculate end video time.")
            end_video_time = None

        return end_video_time

    @property
    def start_time(self):
        if not self._start_time:
            self._start_time = self._get_start_time()
        return self._start_time

    @start_time.setter
    def start_time(self, value: datetime):
        self._start_time = value

    def _get_start_time(self):
        # Get the start time from the video file
        start_time = None
        if self.video_path and os.path.isfile(self.video_path) and self.duration:
            try:
                video = get_video_file_instance(self.video_path)
                start_time = video.start_time + self.duration
            except Exception as e:
                logging.error(f"Failed to get start time: {e}")
        else:
            logging.error("¨Video path and duration are not set. Cannot calculate start time.")

        return start_time

    @property
    def end_time(self):
        if not self._end_time:
            self._end_time = self._get_end_time()
        return self._end_time

    @end_time.setter
    def end_time(self, value: datetime):
        self._end_time = value

    def _get_end_time(self):
        # Get the end time from the video file
        end_time = None
        if self.start_time is not None and self.duration:
            try:
                end_time = self.start_time + self.duration
            except Exception as e:
                logging.error(f"Failed to get end time: {e}")
        else:
            logging.error("Start time and duration are not set. Cannot calculate end time.")

        return end_time

    @property
    def duration(self):
        if not self._duration:
            self._duration = self._get_duration()
        return self._duration

    @duration.setter
    def duration(self, value: timedelta):
        self._duration = value

    def _get_duration(self):
        # Get the duration from the video file
        if self._start_frame is not None and self._end_frame:
            try:
                duration = timedelta(seconds=(self._end_frame - self._start_frame) / self.video_fps)
            except TypeError:
                logging.error("FPS is not set. Cannot calculate duration.")
                duration = None
        elif self._start_video_time is not None and self._end_video_time:
            try:
                duration = self._end_video_time - self._start_video_time
            except TypeError:
                logging.error("Start video time or end video time have incorrect format. Cannot calculate duration.")
                duration = None
        elif self._start_time and self._end_time:
            try:
                duration = self._end_time - self._start_time
            except TypeError:
                logging.error("Start time or end time have incorrect format. Cannot calculate duration.")
                duration = None
        else:
            logging.error("Start frame and end frame are not set. Cannot calculate duration.")
            duration = None

        return duration

    @property
    def video_id(self):
        if not self._video_id:
            self._video_id = self._get_video_id()
        return self._video_id

    @video_id.setter
    def video_id(self, value: str):
        self._video_id = value

    def _get_video_id(self):
        # Get the video ID from the video path
        video_id = None
        if self.video_path:
            video_id = parse_recording_name(self.video_path).get("video_id", None)
        elif self.detection_results:
            if isinstance(self.detection_results, (list, tuple)) and len(self.detection_results) > 0:
                try:
                    video_id = self.detection_results[0].video_id
                except Exception as e:
                    logging.error(f"Failed to get video ID from detection results: {e}")
        else:
            logging.error("Video path is not set. Cannot get video ID.")
        return video_id

    @property
    def recording_id(self):
        if not self._recording_id:
            self._recording_id = self._get_recording_id()
        return self._recording_id

    @recording_id.setter
    def recording_id(self, value: str):
        self._recording_id = value

    def _get_recording_id(self):
        # Get the recording ID from the video path
        recording_id = None
        if self.video_path:
            try:
                recording_id = parse_recording_name(self.video_path).get("recording_id", None)
            except Exception as e:
                logging.error(f"Failed to get recording ID from video path: {e}")
        elif self.video_id:
            try:
                recording_id = os.path.splitext(self.video_id)[0]
            except Exception as e:
                logging.error(f"Failed to get recording ID from video ID: {e}")
        elif self.detection_results:
            if isinstance(self.detection_results, (list, tuple)) and len(self.detection_results) > 0:
                try:
                    recording_id = self.detection_results[0].recording_id
                except Exception as e:
                    logging.error(f"Failed to get recording ID from detection results: {e}")
        else:
            logging.error("Video path is not set. Cannot get recording ID.")
        return recording_id

    @property
    def video_fps(self):
        if not self._video_fps:
            self._video_fps = self._get_video_fps()
        return self._video_fps

    @video_fps.setter
    def video_fps(self, value: int):
        self._video_fps = value

    def _get_video_fps(self):
        # Get the video FPS from the video file
        video_fps = None
        if self.video_path and os.path.isfile(self.video_path):
            try:
                video = get_video_file_instance(self.video_path)
                video_fps = int(video.fps)
            except Exception as e:
                logging.error(f"Failed to get video FPS: {e}")
        elif self.duration and self.start_frame and self._end_frame:
            try:
                video_fps = (self._end_frame - self.start_frame) // self.duration.total_seconds()
            except TypeError:
                logging.error("Start frame, end frame, or duration are not set. Cannot calculate video FPS.")
        else:
            logging.error("Detection results are not set. Cannot get video FPS.")
        return video_fps

    @property
    def flowers(self):
        if not self._flowers:
            self._flowers = self._get_flowers()
        return self._flowers

    @flowers.setter
    def flowers(self, value: int):
        self._flowers = value

    def _get_flowers(self):
        # Get the number of flowers from the detection results
        flowers = None
        if self.detection_results:
            if isinstance(self.detection_results, (list, tuple)) and len(self.detection_results) > 0:
                try:
                    flowers = len(self.detection_results[0].reference_boxes) if self.detection_results[0].reference_boxes else None
                except Exception as e:
                    logging.error(f"Failed to get number of flowers: {e}")
        else:
            logging.error("Detection results are not set. Cannot get number of flowers.")
        return flowers

    @property
    def flowers_boxes(self):
        if not self._flowers_boxes:
            self._flowers_boxes = self._get_flowers_boxes()
        return self._flowers_boxes

    @flowers_boxes.setter
    def flowers_boxes(self, value: Type["DetectionBoxes"]):
        self._flowers_boxes = value

    def _get_flowers_boxes(self):
        # Get the flower boxes from the detection results
        flowers_boxes = None
        if self.detection_results:
            if isinstance(self.detection_results, (list, tuple)) and len(self.detection_results) > 0:
                try:
                    flowers_boxes = self.detection_results[0].reference_boxes if self.detection_results[0].reference_boxes else None
                except Exception as e:
                    logging.error(f"Failed to get flower boxes: {e}")
        else:
            logging.error("Detection results are not set. Cannot get flower boxes.")
        return flowers_boxes

    @property
    def flowers_visited(self):
        if not self._flowers_visited:
            self._flowers_visited = self._get_flowers_visited()
        return self._flowers_visited

    @flowers_visited.setter
    def flowers_visited(self, value: int):
        self._flowers_visited = value

    def _get_flowers_visited(self):
        logging.warning("Method '_get_flowers_visited' is not implemented.")


class AnnotationList(list):
    def __init__(self, *args: Annotation, recording_id: Optional[str] = None):
        super().__init__(args)

        self._recording_id = recording_id
        self._video_paths = []
        self._video_ids = []

    @property
    def recording_id(self):
        if not self._recording_id:
            self._recording_id = self._get_recording_id()
        return self._recording_id

    @recording_id.setter
    def recording_id(self, value: str):
        self._recording_id = value

    def _get_recording_id(self):
        recording_id = None
        if self:
            # Get the recording ID from the annotations by checking if they have the same recording ID
            recording_ids = set([annotation.recording_id for annotation in self])
            if len(recording_ids) > 1:
                logging.warning(f"Annotations have different recording IDs: {recording_ids}")
            else:
                recording_id = recording_ids.pop()
        return recording_id

    @property
    def video_paths(self):
        if not self._video_paths:
            self._video_paths = self._get_video_paths()
        return self._video_paths

    def _get_video_paths(self):
        video_paths = []
        for annotation in self:
            if annotation.video_path:
                video_paths.append(annotation.video_path)
        return video_paths

    @property
    def visits(self):
        return len(self)

    @property
    def detections(self):
        detections = 0
        for annotation in self:
            if annotation.detection_results:
                if isinstance(annotation.detection_results, (list, tuple)):
                    detections += len(annotation.detection_results)
        return detections

    @property
    def visitors(self):
        visitors = 0
        for annotation in self:
            if annotation.detection_results:
                if isinstance(annotation.detection_results, (list, tuple)):
                    try:
                        boxes = [result.boxes for result in annotation.detection_results if result.boxes is not None]
                        visitors += len(set([visitor for box in boxes if box.id is not None for visitor in box.id]))
                    except AttributeError:
                        logging.error("Detection results do not have the attribute boxes. Cannot count visitors.")
        return visitors

    @property
    def detection_results(self):
        detection_results = []
        for annotation in self:
            if annotation.detection_results:
                if isinstance(annotation.detection_results, (list, tuple)):
                    detection_results.extend(annotation.detection_results)
        return detection_results


def align_annotation(visit_start_time: datetime, visit_end_time: datetime, video: Type["Video"]) -> Optional[Annotation]:
    # Check if the visit overlaps with the current video
    try:
        if visit_start_time < video.end_time and visit_end_time > video.start_time:
            annotation_start_time = max(visit_start_time, video.start_time)
            annotation_end_time = min(visit_end_time, video.end_time)
            annotation_duration = annotation_end_time - annotation_start_time

            # Create the Annotation object
            annotation = Annotation(
                video_path=video.video_path,
                start_time=annotation_start_time,
                end_time=annotation_end_time,
                duration=annotation_duration,
                video_fps=video.fps
            )
        else:
            annotation = None
    except Exception as e:
        logging.error(f"Error aligning annotation: {e}")
        annotation = None

    return annotation


def get_annotations_from_dataframe_with_videos(df: pd.DataFrame, video_paths: List[str]) -> AnnotationList:
    # Create Video objects for each video path
    videos = []
    for video_path in video_paths:
        try:
            video = get_video_file_instance(video_path)
        except Exception as e:
            logging.error(f"Error creating video file instance for {video_path}. Error: {e}")
            video = None
        videos.append(video)

    annotations = AnnotationList()

    for _, row in df.iterrows():
        visit_start_time = datetime.strptime(row['ts'], "%Y%m%d_%H_%M_%S")
        visit_duration = timedelta(seconds=row['duration'])
        visit_end_time = visit_start_time + visit_duration

        for video in videos:
            annotation = align_annotation(visit_start_time, visit_end_time, video)

            if annotation:
                annotations.append(annotation)

    return annotations


def get_annotations_from_dataframe(df: pd.DataFrame) -> AnnotationList:
    annotations = AnnotationList()

    for _, row in df.iterrows():
        visit_start_time = datetime.strptime(row['ts'], "%Y%m%d_%H_%M_%S")
        visit_duration = timedelta(seconds=row['duration'])
        visit_end_time = visit_start_time + visit_duration

        annotation = Annotation(
            start_time=visit_start_time,
            end_time=visit_end_time,
            duration=visit_duration
        )

        annotations.append(annotation)

    return annotations


def adjust_annotations_for_videos(annotations: AnnotationList, video_paths: List[str]) -> AnnotationList:
    videos = []
    for video_path in video_paths:
        try:
            video = get_video_file_instance(video_path)
        except Exception as e:
            logging.error(f"Error creating video file instance for {video_path}. Error: {e}")
            video = None
        videos.append(video)

    adjusted_annotations = AnnotationList()

    for annotation in annotations:
        for video in videos:
            adjusted_annotation = align_annotation(annotation.start_time, annotation.end_time, video)

            if adjusted_annotation:
                adjusted_annotations.append(adjusted_annotation)

    return adjusted_annotations

